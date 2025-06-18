import openpyxl
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

# 利用自定义字体解决显示中文问题
myfont = fm.FontProperties(fname="C:\\Windows\\Fonts\\msyh.ttc")    # 微软雅黑
myfont2 = fm.FontProperties(fname="C:\\Windows\\Fonts\\simsun.ttc")    # 宋体
plt.rcParams["font.family"] = myfont.get_name()    # 设置全局字体
# plt.rc('legend', fontsize=10.5)
font_set = {'family' : myfont2.get_name(),
'weight' : 'normal',
'size' : 12,
}

# 从excel中获取每列的数据，使用不同的变量存储，其中前两行不需要读取
# 打开工作簿和表格
workbook = openpyxl.load_workbook('liubianshuju2.xlsx')
sheet = workbook['Sheet1']

# 获取所有行和列
rows = sheet.iter_rows(min_row=4, values_only=True)
# columns = sheet.iter_cols(min_row=4, values_only=True)

# 分别存储每列数据
column1 = []
column2 = []
column3 = []
column4 = []
column5 = []
column6 = []
column7 = []
column8 = []
column9 = []
column10 = []
column11 = []
column12 = []
column13 = []
column14 = []
column15 = []
column16 = []
column17 = []
column18 = []

for row in rows:
    # 将每列数据添加到对应列表中
    column1.append(row[0])
    column2.append(row[1])
    column3.append(row[2])
    column4.append(row[3])
    column5.append(row[4])
    column6.append(row[5])
    column7.append(row[6])
    column8.append(row[7])
    column9.append(row[8])
    column10.append(row[9])
    column11.append(row[10])
    column12.append(row[11])
    column13.append(row[12])
    column14.append(row[13])
    column15.append(row[14])
    column16.append(row[15])
    column17.append(row[16])
    column18.append(row[17])


# 打印数据
print(column1)
print(column2)
print(column3)
print(column4)
print(column18)

# 关闭工作簿
workbook.close()


# 画图
plt.plot(column1, column2, color="#515151", marker="s", label="G"+'\''+" 60mg/mL WT-SH")
plt.plot(column1, column3, color="#515151", marker="^", label="G\" 60mg/mL WT-SH")

plt.plot(column4, column5, color="#F14040", marker="s", label="G"+'\''+" 50mg/mL WT-SH")
plt.plot(column4, column6, color="#F14040", marker="^", label="G\" 50mg/mL WT-SH")

plt.plot(column7, column8, color="#1A6FDF", marker="s", label="G"+'\''+" 40mg/mL WT-SH")
plt.plot(column7, column9, color="#1A6FDF", marker="^", label="G\" 40mg/mL WT-SH")

plt.plot(column10, column11, color="#37AD6B", marker="s", label="G"+'\''+" 30mg/mL WT-SH")
plt.plot(column10, column12, color="#37AD6B", marker="^", label="G\" 30mg/mL WT-SH")

plt.plot(column13, column14, color="#B177DE", marker="s", label="G"+'\''+" 20mg/mL WT-SH")
plt.plot(column13, column15, color="#B177DE", marker="^", label="G\" 20mg/mL WT-SH")

plt.plot(column16, column17, color="#CC9900", marker="s", label="G"+'\''+" 15mg/mL WT-SH")
plt.plot(column16, column18, color="#CC9900", marker="^", label="G\" 15mg/mL WT-SH")

lines = plt.gca().get_lines()

group1 = [0,1]
group2 = [2,3]
group3 = [4,5]
group4 = [6,7]
group5 = [8,9]
group6 = [10,11]

legend1 = plt.legend([lines[i] for i in group1],[lines[i].get_label() for i in group1], loc=(-0.1,-0.45), prop=font_set)
legend2 = plt.legend([lines[i] for i in group2],[lines[i].get_label() for i in group2], loc=(0.30,-0.45), prop=font_set)
legend3 = plt.legend([lines[i] for i in group3],[lines[i].get_label() for i in group3], loc=(0.7,-0.45), prop=font_set)
legend4 = plt.legend([lines[i] for i in group4],[lines[i].get_label() for i in group4], loc=(-0.1,-0.70), prop=font_set)
legend5 = plt.legend([lines[i] for i in group5],[lines[i].get_label() for i in group5], loc=(0.30,-0.70), prop=font_set)
legend6 = plt.legend([lines[i] for i in group6],[lines[i].get_label() for i in group6], loc=(0.7,-0.70), prop=font_set)

plt.gca().add_artist(legend1)
plt.gca().add_artist(legend2)
plt.gca().add_artist(legend3)
plt.gca().add_artist(legend4)
plt.gca().add_artist(legend5)

plt.xscale("log")
plt.yscale("log")
plt.xlabel('剪切应变 γ(%)', font_set)
plt.ylabel('储能模量 G' + '\''+'(Pa)' + '\n' + '损耗模量 G\"'+'(Pa)', font_set)
plt.gcf().subplots_adjust(bottom=0.4)

import time
from datetime import date
current_day = str(date.today())
current_time = time.strftime("%H-%M-%S", time.localtime())
print("当前时间是：", current_day+'-'+current_time)

plt.savefig(current_day+'-'+current_time+'-result.jpg', dpi=600)
# plt.show()

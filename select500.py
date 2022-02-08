"""
从comment.xlsx excel文件中随机选择500条不重复的评论
写入一个新的excel文件中
"""
from openpyxl import load_workbook
from openpyxl import Workbook
import random


def read(filename):
    wb_in = load_workbook(filename=filename, read_only=True)
    sheet_in = wb_in['Sheet1']

    results = set()
    for row in sheet_in.iter_rows():
        for cell in row:
            # print(cell.value)
            results.add(cell.value)
    print(len(results))
    return results


def select(values, nums):
    values_sub = random.sample(values, nums)
    print(len(values_sub))
    print(type(values_sub))
    # print(values_sub)
    return values_sub


def write(values, filename):
    workbook = Workbook()
    save_file = filename
    worksheet = workbook.active
    # 每个workbook创建后，默认会存在一个worksheet，对默认的worksheet进行重命名
    worksheet.title = "Sheet1"
    for row in values:
        row = row.split('qwertyuiop')    # str 2 list 用一个不存在的字符串split，这样就存为一个元素了
        # print(row)
        worksheet.append(row)  # 把每一行append到worksheet中
    workbook.save(filename=save_file)  # 不能忘


if __name__ == "__main__":
    comments = read("comment.xlsx")    # 读入，返回集合类型，去除重复元素
    comments_500 = select(comments, 500)    # 从集合中随机选500个元素，返回列表类型
    write(comments_500, 'comment_500.xlsx')    # 将列表中元素写入新的文件中
